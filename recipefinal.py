import json
import os
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk

import bcrypt
# Required imports
import openpyxl
# Required imports
import pandas as pd
from PIL import Image, ImageTk
from openpyxl.utils import get_column_letter

# Required imports

# Verify login credentials
users_file = "users.json"


def verify_login(username, password):
    users = load_users()
    if username in users and bcrypt.checkpw(password.encode('utf-8'), users[username].encode('utf-8')):
        return True
    return False


# Load users from the JSON file
def load_users():
    if os.path.exists(users_file):
        with open(users_file, 'r', encoding='utf-8') as file:
            return json.load(file)
    return {}


# Add these constants at the top of the file
COLORS = {
    'primary': '#2c3e50',  # Dark blue-gray
    'secondary': '#34495e',  # Lighter blue-gray
    'accent': '#3498db',  # Bright blue
    'success': '#2ecc71',  # Green
    'warning': '#f1c40f',  # Yellow
    'danger': '#e74c3c',  # Red
    'light': '#ecf0f1',  # Light gray
    'dark': '#2c3e50',  # Dark gray
    'white': '#ffffff',  # White
    'tab_bg': '#bdc3c7',  # Tab background
    'tab_selected': '#95a5a6'  # Selected tab
}

FONTS = {
    'header': ('Helvetica', 16, 'bold'),
    'subheader': ('Helvetica', 14, 'bold'),
    'normal': ('Helvetica', 12),
    'small': ('Helvetica', 10),
    'button': ('Helvetica', 11, 'bold')
}


class RecipeManager:
    def __init__(self, root):
        # Initialize main window
        self.root = root
        self.root.withdraw()  # Hide the root window initially
        self.root.title("ManageRecipe Manager")
        self.root.geometry("1200x800")

        # Initialize user-related variables
        self.username = None
        self.recipe_file = "recipes.json"
        self.recipes = []

        # Initialize entry lists
        self.ingredient_entries = []
        self.quantity_entries = []
        self.unit_entries = []
        self.description_entries = []
        self.category_entries = []
        self.search_entry = None
        self.ingredients_dict = {}

        # Initialize UI components
        self.recipe_name_entry = None
        self.image_path_entry = None
        self.recipe_image_label = None
        self.recipe_details_table = None
        self.result_table = None
        self.servings_entry = None
        self.selected_recipes = []
        self.calculated_quantities = None
        self.manager_tasks = None
        self.login_window = None
        self.username_entry = ''
        self.password_entry = ''

        # Initialize search variables
        self.show_search_var = None
        self.update_search_var = None
        self.delete_search_var = None

        # Initialize listboxes
        self.show_listbox = None
        self.update_listbox = None
        self.delete_listbox = None
        self.recipe_table = None
        self.selected_recipes = {}

        # Load recipes_images
        self.load_recipes()
        print(f"Loaded {len(self.recipes)} recipes_images.")

        # Initialize notebook
        self.notebook = None
        self.show_frame = None
        self.add_recipe_frame = None
        self.manage_recipes_frame = None
        self.shopping_recipe_table = None
        self.ingredients_by_category = {}
        self.categories = ["Vegetables", "Fruits", "Frozen", "Not Frozen", "Dairy", "Grains", "Meat", "Others"]

        # Apply theme
        self.style = ttk.Style()
        self.style.configure('Main.TFrame', background=COLORS['light'])
        self.style.configure('Header.TLabel',
                             font=FONTS['header'],
                             background=COLORS['primary'],
                             foreground=COLORS['white'],
                             padding=10,  # Add padding for better appearance
                             relief='flat')  # Flat appearance for modern look

        # Add a new style for section headers
        self.style.configure('SectionHeader.TLabel',
                             font=FONTS['subheader'],
                             background=COLORS['secondary'],
                             foreground=COLORS['white'],
                             padding=5,
                             relief='flat')
        self.style.configure('Success.TButton',
                             font=FONTS['button'],
                             background=COLORS['success'],
                             foreground='black')
        self.style.configure('Warning.TButton',
                             font=FONTS['button'],
                             background=COLORS['warning'])

        # Configure additional styles
        self.style.configure('Card.TLabelframe',
                             background=COLORS['light'],
                             relief='raised')
        self.style.configure('Card.TLabelframe.Label',
                             background=COLORS['light'],
                             foreground=COLORS['dark'],
                             font=FONTS['subheader'])
        self.style.configure('Search.TEntry',
                             fieldbackground=COLORS['white'],
                             foreground=COLORS['dark'])
        self.style.configure('Entry.TEntry',
                             fieldbackground=COLORS['white'],
                             foreground=COLORS['dark'])
        self.style.configure('Normal.TLabel',
                             background=COLORS['light'],
                             foreground=COLORS['dark'],
                             font=FONTS['normal'])
        self.style.configure('Danger.TButton',
                             background=COLORS['danger'],
                             foreground='black',  # Change text color to black
                             font=FONTS['button'])

        # Configure notebook style
        self.style.configure('TNotebook',
                             background=COLORS['tab_bg'])
        self.style.configure('TNotebook.Tab',
                             background=COLORS['tab_bg'],
                             foreground=COLORS['dark'],
                             padding=[10, 5],
                             font=FONTS['normal'])
        self.style.map('TNotebook.Tab',
                       background=[('selected', COLORS['tab_selected'])])

        # Update the Danger.TButton style to ensure text visibility with black text
        self.style.configure('Danger.TButton',
                             background=COLORS['danger'],
                             foreground='black',  # Change text color to black
                             font=FONTS['button'])

        # Add hover state mapping while keeping text black
        self.style.map('Danger.TButton',
                       foreground=[('active', 'black'),  # Keep text black when hovering
                                   ('disabled', COLORS['light'])],  # Light gray when disabled
                       background=[('active', '#c0392b'),  # Darker red when hovering
                                   ('disabled', '#e74c3c')])  # Lighter red when disabled

        # Add new button styles
        self.style.configure('Primary.TButton',
                             font=FONTS['button'],
                             background=COLORS['accent'],
                             foreground=COLORS['white'])

        self.style.configure('Warning.TButton',
                             font=FONTS['button'],
                             background=COLORS['warning'],
                             foreground=COLORS['dark'])

        self.style.configure('Success.TButton',
                             font=FONTS['button'],
                             background=COLORS['success'],
                             foreground='black')

        # Add hover effects for buttons
        self.style.map('Primary.TButton',
                       foreground=[('active', COLORS['white'])],
                       background=[('active', '#2980b9')])  # Darker blue on hover

        self.style.map('Warning.TButton',
                       foreground=[('active', COLORS['dark'])],
                       background=[('active', '#f39c12')])  # Darker yellow on hover

        self.style.map('Success.TButton',
                       foreground=[('active', COLORS['white'])],
                       background=[('active', '#27ae60')])  # Darker green on hover

    def on_login_success(self, username):
        """Handle successful login and set up the GUI."""
        self.username = username
        print(f"Logged in as: {self.username}")
        self.setup_gui()
        # self.root.deiconify()  # Show the main window

    def load_recipes(self):
        """Load recipes_images from a JSON file."""
        try:
            with open(self.recipe_file, "r", encoding="utf-8") as file:
                self.recipes = json.load(file)
                if not isinstance(self.recipes, list):
                    raise ValueError("Invalid format: recipes_images should be a list.")
        except FileNotFoundError:
            print("Recipes file not found. Creating a new one.")
            self.recipes = []
            self.save_recipes()
        except json.JSONDecodeError:
            print("Invalid recipes_images file format. Starting fresh.")
            self.recipes = []
            self.save_recipes()
        except Exception as e:
            print(f"Unexpected error while loading recipes_images: {e}")
            self.recipes = []  # Fallback to an empty list in case of any other error

    def save_recipes(self):
        """Save recipes_images to a JSON file."""
        try:
            with open(self.recipe_file, "w", encoding="utf-8") as file:
                json.dump(self.recipes, file, indent=4)
        except IOError as e:
            messagebox.showerror("Error", f"Failed to save recipes_images: {e}")

    def setup_gui(self):
        """Set up the main GUI with conditional tabs."""
        self.root.deiconify()  # Show the main window

        # Create notebook
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(pady=10, expand=True, fill="both")

        # Create frames for tabs
        self.show_frame = ttk.Frame(self.notebook)
        self.add_recipe_frame = ttk.Frame(self.notebook)
        self.manage_recipes_frame = ttk.Frame(self.notebook)

        # Add basic tabs
        self.setup_show_tab()

        self.notebook.add(self.add_recipe_frame, text="Add ManageRecipe")
        self.setup_add_recipe_tab(self.add_recipe_frame)

        self.setup_shopping_tab()

        # Manager-only tabs
        if self.username == "manager":
            print("User is manager. Adding manager tabs.")
            self.setup_manage_recipes_tab()
            self.setup_manage_users_tab()

    def setup_show_tab(self):
        """Set up the Show Recipes tab with enhanced styling."""
        show_frame = ttk.Frame(self.notebook, style='Main.TFrame')
        self.notebook.add(show_frame, text="Show Recipes")

        # Update header to use new style
        header = ttk.Label(show_frame,
                           text="ManageRecipe Collection",
                           style='Header.TLabel')
        header.pack(fill='x', pady=(0, 10))

        # Update other section headers to use new style
        search_label = ttk.Label(show_frame,
                                 text="Search ManageRecipe:",
                                 style='SectionHeader.TLabel')
        search_label.pack(fill='x', pady=(5, 0))

        # Search frame with styling
        search_frame = ttk.Frame(show_frame, style='Main.TFrame')
        search_frame.pack(pady=5, fill="x")

        ttk.Label(search_frame,
                  text="Search ManageRecipe:",
                  font=FONTS['normal']).pack(side="left", padx=5)
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var)
        search_entry.pack(side="left", padx=5, fill="x", expand=True)

        # ManageRecipe list with styling
        list_frame = ttk.LabelFrame(show_frame,
                                    text="Available Recipes",
                                    style='Subheader.TLabel')
        list_frame.pack(pady=5, fill="both", expand=True)

        self.show_listbox = tk.Listbox(list_frame,
                                       height=10,
                                       font=FONTS['normal'],
                                       bg=COLORS['white'],
                                       selectbackground=COLORS['accent'])
        self.show_listbox.pack(side="left", fill="both", expand=True, padx=5)

        # Scrollbar for recipe list
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.show_listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.show_listbox.configure(yscrollcommand=scrollbar.set)

        # ManageRecipe details
        self.recipe_details = ttk.Frame(show_frame)
        self.recipe_details.pack(pady=10, fill="both", expand=True)

        # Bind events
        self.show_listbox.bind('<<ListboxSelect>>', self.display_recipe)
        self.search_var.trace("w", lambda *args: self.update_show_list(self.search_var.get()))

        # Initial population
        self.update_show_list()

    def update_show_list(self, query=""):
        """Filter and update the recipe list in the Show tab based on the search query."""
        query = query.lower()
        self.show_listbox.delete(0, tk.END)

        if self.recipes:  # Check if recipes_images are loaded
            if query:  # Check if query is not empty
                for recipe in self.recipes:
                    if query in recipe["name"].lower():
                        self.show_listbox.insert(tk.END, recipe["name"])
            else:  # If query is empty, display all recipes_images
                for recipe in self.recipes:
                    self.show_listbox.insert(tk.END, recipe["name"])

        else:
            self.show_listbox.insert(tk.END, "No recipes_images loaded.")  # Informative message

        for recipe in self.recipes:
            if query in recipe["name"].lower():
                self.show_listbox.insert(tk.END, recipe["name"])

    def display_recipe(self, event=None):
        """Display the selected recipe with enhanced styling."""
        for widget in self.recipe_details.winfo_children():
            widget.destroy()

        selection = self.show_listbox.curselection()
        if not selection:
            return

        recipe = self.recipes[selection[0]]

        # Update recipe title styling
        ttk.Label(self.recipe_details,
                  text=f"ManageRecipe: {recipe['name']}",
                  style='Header.TLabel').pack(fill='x', pady=10)

        # Ingredients section
        ingredients_frame = ttk.LabelFrame(self.recipe_details,
                                           text="Ingredients",
                                           style='Subheader.TLabel')
        ingredients_frame.pack(fill='x', pady=5, padx=10)

        for ing in recipe.get("ingredients", []):
            ttk.Label(ingredients_frame,
                      text=f"• {ing['ingredient']} ({ing.get('quantity', '')} {ing.get('unit', '')})",
                      font=FONTS['normal']).pack(pady=2, padx=15)

        # Instructions section
        instructions_frame = ttk.LabelFrame(self.recipe_details,
                                            text="Instructions",
                                            style='Subheader.TLabel')
        instructions_frame.pack(fill='x', pady=5, padx=10)

        for i, inst in enumerate(recipe.get("instructions", []), 1):
            ttk.Label(instructions_frame,
                      text=f"{i}. {inst}",
                      font=FONTS['normal'],
                      wraplength=400).pack(pady=2, padx=15)

        # Add image display
        image_frame = ttk.Frame(self.recipe_details)
        image_frame.pack(pady=10)

        self.recipe_image_label = ttk.Label(image_frame)
        self.recipe_image_label.pack()

        self.image_note_label = ttk.Label(image_frame, text="No image available")
        if "image_path" in recipe and recipe["image_path"] and os.path.exists(recipe["image_path"]):
            try:
                image = Image.open(recipe["image_path"])
                # Resize image if needed
                image = image.resize((200, 200), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(image)
                self.recipe_image_label.configure(image=photo)
                self.recipe_image_label.image = photo
                self.image_note_label.pack_forget()
            except:
                self.recipe_image_label.configure(image="")
                self.image_note_label.pack()
        else:
            self.recipe_image_label.configure(image="")
            self.image_note_label.pack()

    def setup_add_recipe_tab(self, frame):
        """Set up the Add ManageRecipe tab with required input fields."""
        # Set colors
        bg_color = "#f0f0f0"  # Background color
        btn_color = "#4caf50"  # Button color
        btn_text_color = "white"  # Button text color

        # Labels and inputs
        tk.Label(frame, text="ManageRecipe Name:", bg=bg_color).grid(row=0, column=0, sticky="w")
        self.recipe_name_entry = tk.Entry(frame, width=40)
        self.recipe_name_entry.grid(row=0, column=1, padx=5, pady=5)

        # Button to add ingredient rows
        add_row_button = tk.Button(
            frame,
            text="Add Ingredient Row",
            command=lambda: self.add_ingredient_row(frame),
            bg=btn_color,
            fg=btn_text_color
        )
        add_row_button.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="w")

        # Save ManageRecipe Button (Aligned to the right of Add Ingredient)
        save_recipe_button = tk.Button(
            frame,
            text="Save ManageRecipe",
            command=self.save_recipe,
            bg=btn_color,
            fg=btn_text_color
        )
        save_recipe_button.grid(row=1, column=2, columnspan=2, padx=10, pady=10, sticky="w")

        # Image Path
        tk.Label(frame, text="Image Path:", bg=bg_color).grid(row=11, column=0, sticky="w")
        self.image_path_entry = tk.Entry(frame, width=40)
        self.image_path_entry.grid(row=11, column=1, padx=5, pady=5)
        image_path_button = tk.Button(
            frame,
            text="Choose File",
            command=lambda: self.image_path_entry.insert(0, filedialog.askopenfilename())
        )
        image_path_button.grid(row=11, column=2, padx=5, pady=5)

        # Add the first ingredient row on initialization
        self.add_ingredient_row(frame)

    def add_ingredient_row(self, frame):
        """Add a new row for ingredient input."""
        row = len(self.ingredient_entries) + 2

        # Ingredient Name
        tk.Label(frame, text=f"Ingredient {row - 1}:").grid(row=row, column=0, sticky="w")
        ingredient_entry = tk.Entry(frame, width=30)
        ingredient_entry.grid(row=row, column=1, padx=5, pady=5)
        self.ingredient_entries.append(ingredient_entry)

        # Quantity
        tk.Label(frame, text="Quantity:").grid(row=row, column=2, sticky="w")
        quantity_entry = tk.Entry(frame, width=15)
        quantity_entry.grid(row=row, column=3, padx=5, pady=5)
        self.quantity_entries.append(quantity_entry)

        # Unit
        tk.Label(frame, text="Unit:").grid(row=row, column=4, sticky="w")
        unit_entry = ttk.Combobox(
            frame,
            values=["kg", "g", "lb", "oz", "ml", "l", "cup", "tbsp", "tsp", "piece"],
            width=10
        )
        unit_entry.grid(row=row, column=5, padx=5, pady=5)
        self.unit_entries.append(unit_entry)

        # Description
        tk.Label(frame, text="Description:").grid(row=row, column=6, sticky="w")
        description_entry = tk.Entry(frame, width=30)
        description_entry.grid(row=row, column=7, padx=5, pady=5)
        self.description_entries.append(description_entry)

        # Category
        tk.Label(frame, text="Category:").grid(row=row, column=8, sticky="w")
        category_var = tk.StringVar(value="Others")
        category_dropdown = tk.OptionMenu(
            frame,
            category_var,
            "Vegetables", "Fruits", "Frozen Foods", "Not Frozen", "Dairy", "Grains", "Bread", "Meat", "Poultry",
            "Seafood",
            "Vegan",
            "Gluten-Free", "Canned Goods", "Vegetarian", "Organic", "Non-GMO", "Baking Ingredients",
            "Condiments and Sauces", "Beverages", "Sweets", " Snacks", "Oils", " Fats", "Nuts", "Seeds"
                                                                                                "Legumes", "Beans",
            "Others"
        )
        category_dropdown.grid(row=row, column=9, padx=5, pady=5)
        self.category_entries.append(category_var)

    def save_recipe(self):
        """Save the recipe to a file."""
        recipe_name = self.recipe_name_entry.get().strip()
        if not recipe_name:
            messagebox.showerror("Error", "ManageRecipe name is required.")
            return

        # Collect ingredients
        ingredients = [entry.get().strip() for entry in self.ingredient_entries if entry.get().strip()]
        if not ingredients:
            messagebox.showerror("Error", "At least one ingredient is required.")
            return

        # Create the recipe
        recipe = {"name": recipe_name, "ingredients": ingredients, "instructions": []}
        self.recipes.append(recipe)
        self.save_recipes()

        messagebox.showinfo("Success", "ManageRecipe saved successfully.")
        self.recipe_name_entry.delete(0, tk.END)
        for entry in self.ingredient_entries:
            entry.delete(0, tk.END)

    def setup_shopping_tab(self):
        """Set up the 'Shopping List' tab."""

        shopping_frame = ttk.Frame(self.notebook)
        self.notebook.add(shopping_frame, text="Shopping List")

        # Debugging: Check if shopping_frame exists
        print("Initializing shopping_recipe_table...")

        # Status label with Header style and yellow background
        self.status_label = ttk.Label(shopping_frame,
                                      text="Select recipes_images and set servings to calculate shopping list",
                                      style='Header.TLabel',
                                      background=COLORS['warning'],  # Yellow background
                                      foreground='black')  # Black text
        self.status_label.pack(pady=5)

        # Left side - ManageRecipe selection
        left_frame = ttk.Frame(shopping_frame)
        left_frame.pack(side="left", fill="both", expand=True, padx=5, pady=5)

        # ManageRecipe selection area with bold header and yellow background
        selection_frame = ttk.LabelFrame(left_frame,
                                         text="Select Recipes",
                                         style='Warning.TLabelframe')  # New style for yellow background
        selection_frame.pack(fill="x", padx=5, pady=5)

        # Debugging: Check if selection_frame exists
        print("selection_frame initialized:", selection_frame)

        # Search functionality with styled label
        search_frame = ttk.Frame(selection_frame)
        search_frame.pack(fill="x", padx=5, pady=5)

        ttk.Label(search_frame,
                  text="Search ManageRecipe:",
                  style='Subheader.TLabel').pack(side='left')

        self.shopping_search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame,
                                 textvariable=self.shopping_search_var,
                                 font=("Arial", 11))
        search_entry.pack(side='left', expand=True, fill='x', padx=5)

        # Bind search entry to update function
        self.shopping_search_var.trace("w", lambda *args: self.update_shopping_table())

        # ManageRecipe table
        columns = ("ManageRecipe", "Servings")
        self.recipe_table = ttk.Treeview(selection_frame, columns=columns, show="headings", height=6,
                                         selectmode="extended")
        self.recipe_table.heading("ManageRecipe", text="ManageRecipe Name")
        self.recipe_table.heading("Servings", text="Servings")
        self.recipe_table.pack(fill="x", padx=5, pady=5)

        # Populate the recipe table with available recipes_images
        for recipe in self.recipes:
            self.recipe_table.insert('', 'end', values=(recipe['name'], '0'))

        # Debugging: Check after inserting recipes_images
        print("Treeview populated with recipes_images.")

        # Bind double-click to edit servings
        # self.shopping_recipe_table.bind('<Double-1>', self.edit_servings)

        # Servings input with styled label
        servings_frame = ttk.Frame(left_frame)
        servings_frame.pack(fill="x", padx=5, pady=5)

        ttk.Label(servings_frame,
                  text="Number of Servings:",
                  style='Subheader.TLabel').pack(side="left")

        self.servings_entry = ttk.Entry(servings_frame, width=10)
        self.servings_entry.pack(side="left", padx=5)
        self.servings_entry.insert(0, "1")

        # Bind <Return> key to update servings
        self.servings_entry.bind("<Return>", self.update_servings)

        # Apply Servings button with Warning style for black text
        ttk.Button(servings_frame,
                   text="Apply Servings",
                   style='Warning.TButton',
                   command=self.apply_servings).pack(side="left", padx=5)

        # Control buttons with appropriate styles
        button_frame = ttk.Frame(left_frame)
        button_frame.pack(fill="x", padx=5, pady=5)

        # Button instances with commands
        ttk.Button(button_frame,
                   text="Calculate",
                   command=self.calculate_shopping_list
                   ).pack(side="left", padx=5)

        ttk.Button(button_frame,
                   text="Reset",
                   command=self.reset_shopping_list
                   ).pack(side="left", padx=5)

        ttk.Button(button_frame,
                   text="Export to Excel",
                   command=self.export_to_excel_action
                   ).pack(side="left", padx=5)

        # Right side - Results with styled header
        right_frame = ttk.Frame(shopping_frame)
        right_frame.pack(side="right", fill="both", expand=True, padx=5, pady=5)

        # Shopping List header with section header style
        ttk.Label(right_frame,
                  text="Shopping List:",
                  style='SectionHeader.TLabel').pack(fill="x")

        # Results table
        columns = ("Ingredient", "Quantity", "Unit", "Category", "Description")
        self.result_table = ttk.Treeview(right_frame, columns=columns, show="headings", height=10)

        for col in columns:
            self.result_table.heading(col, text=col)
            self.result_table.column(col, width=100)

        self.result_table.pack(fill="both", expand=True)

        # Add feedback label
        self.servings_feedback = ttk.Label(servings_frame,
                                           text="",
                                           style='Normal.TLabel')
        self.servings_feedback.pack(side="left", padx=5)

    def update_shopping_table(self):
        """Update the recipe table in the 'Shopping List' tab based on the search term."""
        search_term = self.shopping_search_var.get().lower()

        # Clear existing items in the table
        self.recipe_table.delete(*self.recipe_table.get_children())

        # Temporarily store selected recipes_images and their servings
        temp_selected_recipes = self.selected_recipes.copy()

        # Filter recipes_images based on search term
        for recipe in self.recipes:
            if search_term in recipe['name'].lower() or not search_term:
                servings = temp_selected_recipes.get(recipe['name'], '0')
                self.recipe_table.insert('', 'end', values=(recipe['name'], servings))

        # Update selected_recipes based on current table selections
        self.selected_recipes = {}
        for item_id in self.recipe_table.selection():
            recipe_name = self.recipe_table.item(item_id)['values'][0]
            servings = int(self.recipe_table.item(item_id)['values'][1])
            self.selected_recipes[recipe_name] = servings

    def update_servings(self, event):
        """Update servings for the selected recipe in the shopping list."""
        selected_item = self.recipe_table.selection()
        if selected_item:
            servings = self.servings_entry.get()
            try:
                servings = int(servings)
                if servings <= 0:
                    raise ValueError
                self.recipe_table.set(selected_item, "Servings", servings)
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid positive number for servings.")

    def filter_shopping_recipes(self):
        """Filter recipes_images based on search term while preserving servings values"""
        search_term = self.shopping_search_var.get().lower()

        # Store current servings values before clearing
        current_servings = {}
        for item in self.recipe_table.get_children():
            recipe_name = self.recipe_table.item(item)['values'][0]
            servings = self.recipe_table.item(item)['values'][1]
            current_servings[recipe_name] = servings

        # Clear current items
        for item in self.recipe_table.get_children():
            self.recipe_table.delete(item)

        # Add filtered recipes_images with preserved servings
        for recipe in self.recipes:
            if search_term in recipe['name'].lower() or not search_term:
                # Get the previously set servings value or default to '0'
                servings = current_servings.get(recipe['name'], '0')
                self.recipe_table.insert('', 'end', values=(recipe['name'], servings))

    def apply_servings(self):
        """Apply the entered servings to selected recipes_images."""
        selection = self.recipe_table.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select at least one recipe")
            return

        try:
            servings = float(self.servings_entry.get())
            if servings <= 0:
                raise ValueError("Servings must be positive")

            # Update servings for selected recipes_images
            for item in selection:
                self.recipe_table.set(item, "Servings", str(servings))

            self.servings_feedback.config(
                text=f"✓ Updated servings to {servings}",
                foreground="green")

        except ValueError as e:
            self.servings_feedback.config(
                text=" Enter a valid positive number",
                foreground="red")

    def calculate_shopping_list_old(self):
        """Calculate ingredients based on selected recipes_images and servings"""
        selection = self.recipe_table.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select at least one recipe")
            return

        # Clear previous results
        for item in self.result_table.get_children():
            self.result_table.delete(item)

        for row in self.recipe_table.get_children():
            recipe_name, servings = self.recipe_table.item(row, "values")
            print(f"servings is {servings} type of {type(servings)}")

            servings = float(servings)

            # Skip if servings is 0
            if servings <= 0:
                continue

            recipe = next((r for r in self.recipes if r['name'] == recipe_name), None)
            if recipe:
                for ing in recipe.get('ingredients', []):
                    key = (ing.get('ingredient', ''), ing.get('unit', ''))
                    try:
                        amount = float(ing.get('quantity', 0)) * servings
                    except (ValueError, TypeError):
                        amount = 0

                    if key in self.ingredients_dict:
                        self.ingredients_dict[key]['amount'] += amount
                    else:
                        self.ingredients_dict[key] = {
                            'amount': amount,
                            'category': ing.get('category', ''),
                            'description': ing.get('description', '')
                        }

        # Display results and  Group ingredients by category
        self.ingredients_by_category = {}
        for (ingredient, unit), data in self.ingredients_dict.items():
            if data['amount'] > 0:  # Only show ingredients with non-zero amounts
                category = data['category']
                if category not in self.ingredients_by_category:
                    self.ingredients_by_category[category] = []
                self.ingredients_by_category[category].append((ingredient, unit, data['amount'], data['description']))

                self.result_table.insert('', 'end', values=(
                    ingredient,
                    f"{data['amount']:.2f}",
                    unit,
                    data['category'],
                    data['description']
                ))

        ingredients_df = pd.DataFrame(self.ingredients_dict,
                                      columns=["ingredient", "quantity", "unit", "category", "description"])
        print(ingredients_df)

        if self.result_table.get_children():
            self.status_label.config(text="Shopping list calculated successfully")

        else:
            self.status_label.config(text="No ingredients to display. Please check recipe servings.")

    def calculate_shopping_list(self):
        """Calculate ingredients based on selected recipes_images and servings"""
        selection = self.recipe_table.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select at least one recipe")
            return

        # Clear previous results
        for item in self.result_table.get_children():
            self.result_table.delete(item)

        for row in self.recipe_table.get_children():
            recipe_name, servings = self.recipe_table.item(row, "values")
            print(f"servings is {servings} type of {type(servings)}")

            servings = float(servings)

            # Skip if servings is 0
            if servings <= 0:
                continue
            recipe = {}
            print(recipe, "  ", servings)
            recipe = next((r for r in self.recipes if r['name'] == recipe_name), None)
            print(recipe)

            if recipe:
                for ing in recipe.get('ingredients', []):
                    key = ing.get('ingredient', '')
                    try:
                        amount = float(ing.get('quantity', 0)) * servings
                    except (ValueError, TypeError):
                        amount = 0

                    if key in self.ingredients_dict:
                        self.ingredients_dict[key]['amount'] += amount
                        self.ingredients_dict[key] = {
                            'ingredient': key,
                            'amount': amount,
                            'unit': ing.get('unit', ''),
                            'category': ing.get('category', ''),
                            'description': ing.get('description', '')
                        }


                    else:
                        self.ingredients_dict[key] = {
                            'ingredient': key,
                            'amount': amount,
                            'unit': ing.get('unit', ''),
                            'category': ing.get('category', ''),
                            'description': ing.get('description', '')
                        }

        ingredients_data = []
        for ingredient, details in self.ingredients_dict.items():
            ingredients_data.append(
                [ingredient, details['amount'], details['unit'], details['category'], details['description']])

        ingredients_df = pd.DataFrame(ingredients_data,
                                      columns=['ingredient', 'quantity', 'unit', 'category', 'description'])
        print(ingredients_df)

        ingredients_df = ingredients_df.sort_values(by=['category', 'ingredient'])
        print(ingredients_df)

        # Populate the result table
        self.result_table.delete(*self.result_table.get_children())

        row_num = 1
        for index, row in ingredients_df.iterrows():
            ingredient, unit, amount, category, description = row.values
            print(ingredient, " ", unit, " ", amount, " ", category, " ", description)

            self.result_table.insert('', 'end', values=(
                ingredient,
                amount,
                unit,
                category,
                description
            ))
            row_num += 1

        if self.result_table.get_children():
            self.status_label.config(text="Shopping list calculated successfully")
            return ingredients_df
        else:
            self.status_label.config(text="No ingredients to display. Please check recipe servings.")

    def calculate_ingredients(self, selected_recipes):
        # Calculate ingredients based on selected recipes_images and servings
        ingredients_dict = {}
        for recipe_name, servings in selected_recipes.items():
            recipe = next((r for r in self.recipes if r['name'] == recipe_name), None)
            if recipe:
                for ingredient in recipe['ingredients']:
                    ingredient_name = ingredient['ingredient']
                    unit = ingredient['unit']
                    quantity = float(ingredient['quantity']) * servings

                    if (ingredient_name, unit) in ingredients_dict:
                        ingredients_dict[(ingredient_name, unit)]['amount'] += quantity
                    else:
                        ingredients_dict[(ingredient_name, unit)] = {
                            'amount': quantity,
                            'category': ingredient['category'],
                            'description': ingredient['description']
                        }
        return ingredients_dict

    def reset_shopping_list(self):
        """Reset the shopping list and clear selections."""
        # Clear recipe selections
        self.recipe_table.selection_remove(self.recipe_table.selection())

        # Reset all servings to 0
        for item in self.recipe_table.get_children():
            self.recipe_table.set(item, "Servings", "0")

        # Clear result table
        for item in self.result_table.get_children():
            self.result_table.delete(item)

        # Reset servings entry
        self.servings_entry.delete(0, tk.END)
        self.servings_entry.insert(0, "1")

        # Clear servings feedback
        self.servings_feedback.config(text="")

        # Update status
        self.status_label.config(text="Shopping list reset")

    def setup_manage_recipes_tab(self):
        """Set up the 'Manage Recipes' tab with Update and Delete sub-tabs."""
        manage_frame = ttk.Frame(self.notebook)
        self.notebook.add(manage_frame, text="Manage Recipes")

        # Create sub-notebook for Update and Delete tabs
        sub_notebook = ttk.Notebook(manage_frame)
        sub_notebook.pack(fill="both", expand=True, padx=5, pady=5)

        # Create Update and Delete frames
        self.update_frame = ttk.Frame(sub_notebook)
        self.delete_frame = ttk.Frame(sub_notebook)

        # Add frames to sub-notebook
        sub_notebook.add(self.update_frame, text="Update ManageRecipe")
        sub_notebook.add(self.delete_frame, text="Delete ManageRecipe")

        # Setup Update and Delete tabs
        self.setup_update_tab()
        self.setup_delete_tab()

    def delete_recipe(self):
        """Delete the selected recipe."""
        selection = self.delete_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a recipe to delete")
            return

        recipe_name = self.delete_listbox.get(selection[0])
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete '{recipe_name}'?"):
            # Remove recipe from list
            self.recipes = [r for r in self.recipes if r['name'] != recipe_name]

            # Save changes
            try:
                self.save_recipes()
                messagebox.showinfo("Success", "ManageRecipe deleted successfully!")
                self.refresh_lists()  # Refresh all recipe lists
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete recipe: {str(e)}")

    def update_delete_list(self):
        """Update the delete listbox with filtered recipes_images."""
        self.delete_listbox.delete(0, tk.END)
        search_term = self.delete_search_var.get().lower() if hasattr(self, 'delete_search_var') else ""

        for recipe in self.recipes:
            if search_term in recipe['name'].lower():
                self.delete_listbox.insert(tk.END, recipe['name'])

    def add_ingredient_row_update(self, frame):
        """Add a new ingredient row in the update form."""
        row = len(self.ingredient_vars) + 1
        row_vars = []

        for j, field in enumerate(["ingredient", "category", "quantity", "unit", "description"]):
            var = tk.StringVar()
            entry = ttk.Entry(frame, textvariable=var)
            entry.grid(row=row, column=j, padx=2, pady=2)
            row_vars.append(var)

        self.ingredient_vars.append(row_vars)

    def setup_manage_users_tab(self):
        """Set up the 'Manage Users' tab for managers."""
        manage_users_frame = ttk.Frame(self.notebook, style='Main.TFrame')
        self.notebook.add(manage_users_frame, text="Manage Users")

        # Create main container with background color
        container = ttk.Frame(manage_users_frame, style='Main.TFrame')
        container.pack(fill="both", expand=True, padx=10, pady=10)

        # Add search functionality
        search_frame = ttk.Frame(container, style='Main.TFrame')
        search_frame.pack(fill="x", pady=5)

        ttk.Label(search_frame,
                  text="Search Users:",
                  style='Subheader.TLabel').pack(side="left", padx=5)

        self.users_search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame,
                                 textvariable=self.users_search_var,
                                 style='Search.TEntry')
        search_entry.pack(side="left", fill="x", expand=True, padx=5)

        # Left side - User list with colored background
        left_frame = ttk.LabelFrame(container,
                                    text="User List",
                                    style='Card.TLabelframe')
        left_frame.pack(side="left", fill="both", expand=True, padx=5, pady=5)

        # User listbox with scrollbar and colors
        self.users_listbox = tk.Listbox(left_frame,
                                        height=15,
                                        bg=COLORS['white'],
                                        fg=COLORS['dark'],
                                        selectbackground=COLORS['accent'],
                                        selectforeground=COLORS['white'])
        scrollbar = ttk.Scrollbar(left_frame, orient="vertical",
                                  command=self.users_listbox.yview)

        self.users_listbox.configure(yscrollcommand=scrollbar.set)
        self.users_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Right side - User management with colored buttons
        right_frame = ttk.LabelFrame(container,
                                     text="User Management",
                                     style='Card.TLabelframe')
        right_frame.pack(side="right", fill="both", expand=True, padx=5, pady=5)

        # Add user section with styled components
        add_frame = ttk.Frame(right_frame, style='Main.TFrame')
        add_frame.pack(fill="x", pady=10)

        ttk.Label(add_frame,
                  text="Add New User",
                  style='Subheader.TLabel').pack()

        # Username entry
        username_frame = ttk.Frame(add_frame, style='Main.TFrame')
        username_frame.pack(fill="x", pady=5)
        ttk.Label(username_frame,
                  text="Username:",
                  style='Normal.TLabel').pack(side="left")
        self.new_username = ttk.Entry(username_frame, style='Entry.TEntry')
        self.new_username.pack(side="left", padx=5)

        # Password entry
        password_frame = ttk.Frame(add_frame, style='Main.TFrame')
        password_frame.pack(fill="x", pady=5)
        ttk.Label(password_frame,
                  text="Password:",
                  style='Normal.TLabel').pack(side="left")
        self.new_password = ttk.Entry(password_frame,
                                      show="*",
                                      style='Entry.TEntry')
        self.new_password.pack(side="left", padx=5)

        # Action buttons with colors
        button_frame = ttk.Frame(right_frame, style='Main.TFrame')
        button_frame.pack(fill="x", pady=10)

        ttk.Button(button_frame,
                   text="Add User",
                   style='Success.TButton',
                   command=self.add_user).pack(pady=5)

        delete_button = ttk.Button(button_frame,
                                   text="Delete Selected User",
                                   style='Danger.TButton',
                                   command=self.delete_user)
        delete_button.pack(pady=5)

        # Bind search functionality
        self.users_search_var.trace("w", lambda *args: self.filter_users())

        # Initial population of the list
        self.refresh_user_list()

    def filter_users(self):
        """Filter users based on search term."""
        search_term = self.users_search_var.get().lower()
        self.users_listbox.delete(0, tk.END)

        users = load_users()
        for username in users.keys():
            if search_term in username.lower():
                self.users_listbox.insert(tk.END, username)

    def refresh_user_list(self):
        """Refresh the list of users in the manage users tab."""
        if hasattr(self, 'users_listbox'):
            self.users_listbox.delete(0, tk.END)
            users = load_users()
            for username in users.keys():
                self.users_listbox.insert(tk.END, username)

    def add_user(self):
        """Add a new user to the system."""
        username = self.new_username.get().strip()
        password = self.new_password.get().strip()

        if not username or not password:
            messagebox.showerror("Error", "Username and password are required")
            return

        users = load_users()
        if username in users:
            messagebox.showerror("Error", "Username already exists")
            return

        # Hash the password
        hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
        users[username] = hashed.decode('utf-8')

        # Save to file
        try:
            with open(users_file, 'w', encoding='utf-8') as f:
                json.dump(users, f)
            messagebox.showinfo("Success", "User added successfully")
            self.new_username.delete(0, tk.END)
            self.new_password.delete(0, tk.END)
            self.refresh_user_list()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add user: {str(e)}")

    def delete_user(self):
        """Delete the selected user."""
        selection = self.users_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a user to delete")
            return

        username = self.users_listbox.get(selection[0])
        if username == "manager":
            messagebox.showerror("Error", "Cannot delete manager account")
            return

        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete user '{username}'?"):
            users = load_users()
            if username in users:
                del users[username]
                try:
                    with open(users_file, 'w', encoding='utf-8') as f:
                        json.dump(users, f)
                    messagebox.showinfo("Success", "User deleted successfully")
                    self.refresh_user_list()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to delete user: {str(e)}")

    def export_to_excel_action(self):
        try:
            ingredients_df = self.calculate_shopping_list()
        except Exception as e:
            messagebox.showerror("Error", f"Error calculating shopping list: {str(e)}")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not file_path:
            return

        # Group by category and ingredient, sum quantities, and create new DataFrame
        df_combined = ingredients_df.groupby(['category', 'ingredient']).agg(
            {'quantity': 'sum', 'unit': 'first'}).reset_index()

        # Convert quantities to practical units
        df_combined['recipe_amount'], df_combined['package_recommendation'] = zip(
            *df_combined.apply(lambda x: self.convert_to_practical_units(x['quantity'], x['unit'], x['ingredient']),
                               axis=1))

        # Sort by category and ingredient
        df_combined = df_combined.sort_values(by=['category', 'ingredient'])

        # Create Excel workbook with styling
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Shopping List"

        # Styles
        header_font = openpyxl.styles.Font(bold=True)
        header_fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        light_blue_fill = openpyxl.styles.PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                                                      fill_type="solid")  # Light blue color
        white_fill = openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Write headers
        headers = ["Category", "Ingredient", "Amount"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill

            # Write data with alternating colors
            for row_idx, row in enumerate(df_combined.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = str(value)
                    cell.fill = light_blue_fill if row_idx % 2 == 0 else white_fill

        # Adjust column widths
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        # Save workbook
        wb.save(file_path)
        wb.close()
        messagebox.showinfo("Success", f"Shopping list exported to {file_path}")

    def sort_recipes_alphabetically(self, json_file='recipes_images.json'):
        """Sort recipes_images in the JSON file alphabetically by the recipe name."""
        # Load existing recipes_images from the JSON file
        if os.path.exists(json_file):
            with open(json_file, 'r', encoding='utf-8') as file:
                recipes = json.load(file)
        else:
            print(f"{json_file} not found!")
            return

        # Sort recipes_images by name
        sorted_recipes = sorted(recipes, key=lambda x: x['name'].lower())  # Alphabetically by 'name'

        # Save sorted recipes_images back to the JSON file
        with open(json_file, 'w', encoding='utf-8') as file:
            json.dump(sorted_recipes, file, indent=4)

        print(f"Recipes have been sorted and saved to {json_file}.")
        # change 24/12/2024 adding return
        return sorted_recipes

    def edit_servings(self, event):
        """Enable editing of servings directly in the Treeview."""
        # Identify the row and column clicked
        item = self.shopping_recipe_table.selection()[0]  # Get the selected row
        column = self.shopping_recipe_table.identify_column(event.x)  # Get the clicked column

        if column == "#2":  # Only allow editing of the "Servings" column
            # Get the current bounding box of the cell
            bbox = self.shopping_recipe_table.bbox(item, "Servings")
            x, y, width, height = bbox

            # Create an entry widget over the cell
            servings_entry = ttk.Entry(self.shopping_recipe_table, font=("Arial", 10))
            servings_entry.place(x=x, y=y, width=width, height=height)
            servings_entry.insert(0, self.shopping_recipe_table.set(item, "Servings"))

            def save_servings(event):
                """Save the edited value back to the Treeview."""
                new_value = servings_entry.get()
                try:
                    # Validate input as a positive integer
                    new_value = int(new_value)
                    if new_value < 0:
                        raise ValueError
                except ValueError:
                    messagebox.showerror("Invalid Input", "Servings must be a positive integer.")
                    servings_entry.destroy()
                    return

                # Update the Treeview with the new value
                self.shopping_recipe_table.set(item, "Servings", str(new_value))
                servings_entry.destroy()

            servings_entry.bind("<Return>", save_servings)  # Save on Enter key
            servings_entry.focus_set()

    def calculate(self):
        # Clear the result table
        for row in self.result_table.get_children():
            self.result_table.delete(row)

        # Example calculation logic
        total_ingredients = {}

        # Iterate through selected recipes_images
        for item in self.recipe_table.selection():
            recipe_name = self.recipe_table.item(item, "values")[0]

            # Load recipes_images from JSON
            try:
                with open("recipes.json", "r") as file:
                    recipes = json.load(file)
            except (FileNotFoundError, json.JSONDecodeError):
                recipes = []

            # Find the selected recipe
            recipe = next((r for r in recipes if r["name"] == recipe_name), None)
            if recipe:
                for ingredient in recipe["ingredients"]:
                    key = (ingredient["ingredient"], ingredient["unit"])
                    quantity = float(ingredient["quantity"]) if isinstance(ingredient["quantity"], (int, float)) else 0
                    total_ingredients[key] = total_ingredients.get(key, 0) + quantity

        # Populate result table with aggregated quantities
        for (ingredient, unit), quantity in total_ingredients.items():
            self.result_table.insert("", tk.END, values=(ingredient, f"{quantity} {unit}"))

    # ... other methods for displaying recipes_images, filtering, etc.
    def convert_to_practical_units_old(self, amount, unit, ingredient_type):
        """
        Convert measurements to practical shopping units with minimum retail packaging recommendations.

        Args:
            amount (float): The amount to convert
            unit (str): The unit of measurement
            ingredient_type (str): Type of ingredient

        Returns:
            tuple: (recipe amount, recommended package size)
        """

        # Detailed retail packaging information
        retail_sizes = {
            # Baking & Pantry
            'rice': {
                'minimum_sizes': {
                    'cup': '10 lb bag',
                    'pound': '5 lb bag'
                },
                'conversions': {
                    'cup': 0.5  # pounds per cup
                }
            },

            'pasta': {
                'minimum_sizes': {
                    'pound': '1 lb box',
                    'kilogram': '2 kg bag'
                },
                'conversions': {
                    'pound': 1  # pounds per pound
                }
            },

            'canned_tomatoes': {
                'minimum_sizes': {
                    'can': '14.5 oz can'
                },
                'conversions': {
                    'can': 14.5  # ounces per can
                }
            },
            'sugar': {
                'minimum_sizes': {
                    'teaspoon': '1 lb bag',
                    'tablespoon': '1 lb bag',
                    'cup': '2 lb bag',
                    'pound': '1 lb bag'
                },
                'conversions': {
                    'teaspoon': 0.0104167,  # pounds
                    'tablespoon': 0.0312500,  # pounds
                    'cup': 0.5,  # pounds
                }
            },
            'granulated_sugar': {
                'minimum_sizes': {
                    'teaspoon': '1 lb bag',
                    'tablespoon': '1 lb bag',
                    'cup': '2 lb bag',
                    'pound': '1 lb bag'
                },
                'conversions': {'cup': 0.5}  # pounds per cup
            },
            'brown_sugar': {
                'minimum_sizes': {
                    'teaspoon': '1 lb box',
                    'tablespoon': '1 lb box',
                    'cup': '2 lb box',
                    'pound': '1 lb box'
                },
                'conversions': {'cup': 0.5}  # pounds per cup
            },
            'powdered_sugar': {
                'minimum_sizes': {
                    'teaspoon': '1 lb box',
                    'tablespoon': '1 lb box',
                    'cup': '2 lb box',
                    'pound': '1 lb box'
                },
                'conversions': {'cup': 0.5}  # pounds per cup
            },
            'raisins': {
                'minimum_sizes': {
                    'teaspoon': '1.5 oz box',
                    'tablespoon': '1.5 oz box',
                    'cup': '15 oz box'
                },
                'conversions': {'cup': 5}  # ounces per cup
            },
            'all_purpose_flour': {
                'minimum_sizes': {
                    'teaspoon': '2 lb bag',
                    'tablespoon': '2 lb bag',
                    'cup': '5 lb bag',
                    'pound': '2 lb bag'
                },
                'conversions': {'cup': 0.4125}  # pounds per cup
            },
            'flour': {
                'minimum_sizes': {
                    'teaspoon': '1 lb bag',
                    'tablespoon': '1 lb bag',
                    'cup': '2 lb bag',
                    'pound': '1 lb bag'
                },
                'conversions': {
                    'teaspoon': 0.00859375,  # pounds
                    'tablespoon': 0.025781,  # pounds
                    'cup': 0.4125,  # pounds
                }
            },
            # Liquids
            'vegetable_broth': {
                'minimum_sizes': {
                    'teaspoon': '8 fl oz carton',
                    'tablespoon': '8 fl oz carton',
                    'cup': '32 fl oz carton',
                    'fluid_ounce': '8 fl oz carton'
                },
                'conversions': {
                    'teaspoon': 0.166667,  # fluid ounces
                    'tablespoon': 0.5,  # fluid ounces
                    'cup': 8,  # fluid ounces
                }
            },
            'olive_oil': {
                'minimum_sizes': {
                    'teaspoon': '8.5 fl oz bottle',
                    'tablespoon': '8.5 fl oz bottle',
                    'cup': '16 fl oz bottle',
                    'fluid_ounce': '8.5 fl oz bottle'
                },
                'conversions': {
                    'teaspoon': 0.166667,  # fluid ounces
                    'tablespoon': 0.5,  # fluid ounces
                    'cup': 8,  # fluid ounces
                }
            },
            # Spices & Seasonings
            'ground_cinnamon': {
                'minimum_sizes': {
                    'teaspoon': '1.5 oz jar',
                    'tablespoon': '1.5 oz jar',
                    'cup': '2.37 oz jar',
                    'ounce': '1.5 oz jar'
                },
                'conversions': {
                    'teaspoon': 0.0729167,  # ounces
                    'tablespoon': 0.21875,  # ounces
                    'cup': 3.5,  # ounces
                }
            },
            'vanilla_extract': {
                'minimum_sizes': {
                    'teaspoon': '1 fl oz bottle',
                    'tablespoon': '2 fl oz bottle',
                    'cup': '4 fl oz bottle',
                    'fluid_ounce': '1 fl oz bottle'
                },
                'conversions': {
                    'teaspoon': 0.166667,  # fluid ounces
                    'tablespoon': 0.5,  # fluid ounces
                    'cup': 8,  # fluid ounces
                }
            },
            # Dairy & Refrigerated
            'milk': {
                'minimum_sizes': {
                    'teaspoon': '8 fl oz bottle',
                    'tablespoon': '8 fl oz bottle',
                    'cup': '16 fl oz bottle',
                    'fluid_ounce': '8 fl oz bottle'
                },
                'conversions': {
                    'teaspoon': 0.166667,  # fluid ounces
                    'tablespoon': 0.5,  # fluid ounces
                    'cup': 8,  # fluid ounces
                }
            },
            'baking_powder': {
                'minimum_sizes': {
                    'teaspoon': '8 oz can',
                    'tablespoon': '8 oz can',
                    'cup': '8 oz can'
                },
                'conversions': {'cup': 8}  # ounces per cup
            },

            # Condiments
            'garlic': {
                'minimum_sizes': {
                    'clove': '1 bulb',
                },
                'conversions': {'clove': 1}  # Each clove is a single unit
            },
            'lemon_juice': {
                'minimum_sizes': {
                    'teaspoon': '4 fl oz bottle',
                    'tablespoon': '8 fl oz bottle',
                    'cup': '16 fl oz bottle'
                },
                'conversions': {'cup': 8}  # fluid ounces per cup
            },

            # Dairy
            'butter': {
                'minimum_sizes': {
                    'gram': '200g package',
                    'cup': '16 oz package (2 cups)',
                    'pound': '1 lb package'
                },
                'conversions': {'gram': 0.0353}  # pounds per gram
            },
            # Sweets
            'dark_chocolate': {
                'minimum_sizes': {
                    'gram': '200g bar',
                    'ounce': '7 oz bar',
                    'pound': '1 lb bag'
                },
                'conversions': {'gram': 0.0353}  # pounds per gram
            }
        }

        try:
            amount = float(amount)
        except (ValueError, TypeError):
            return f"{amount} {unit}", "Unable to determine package size"

        ingredient = ingredient_type.lower().strip()
        if ingredient in retail_sizes:
            data = retail_sizes[ingredient]
            conversion_factor = data['conversions'].get(unit, 1)  # Default to 1 if no conversion is defined
            converted_amount = amount * conversion_factor

            # Get recommended package size based on converted amount
            recommended_package = None
            for package_size, details in data['minimum_sizes'].items():
                if isinstance(details, dict) and converted_amount >= details.get('min_amount', 0):
                    recommended_package = details
                    break
            if recommended_package:
                # Ensure consistent unit for package size recommendation (e.g., ounces)
                if 'unit' in recommended_package:
                    package_unit = recommended_package['unit']
                    if package_unit != unit and package_unit in data['conversions']:
                        package_amount = converted_amount / float(data['conversions'][package_unit])
                        return f"{package_amount:.2f} {package_unit}", recommended_package['description']
                return converted_amount, recommended_package['description']
            else:
                return converted_amount, "N/A (consider smaller unit)"  # Suggest using a smaller unit

        return f"{amount} {unit}", "Unable to determine package size"

    def convert_to_practical_units(self, amount, unit, ingredient_type):
        """
               Convert measurements to practical shopping units with minimum retail packaging recommendations.

               Args:
                   amount (float): The amount to convert
                   unit (str): The unit of measurement
                   ingredient_type (str): Type of ingredient

               Returns:
                   tuple: (recipe amount, recommended package size)
               """

        # Detailed retail packaging information
        retail_sizes = {
            # Baking & Pantry
            'rice': {
                'minimum_sizes': {
                    'cup': '10 lb bag',
                    'pound': '5 lb bag'
                },
                'conversions': {
                    'cup': 0.5  # pounds per cup
                }
            },

            'pasta': {
                'minimum_sizes': {
                    'pound': '1 lb box',
                    'kilogram': '2 kg bag'
                },
                'conversions': {
                    'pound': 1  # pounds per pound
                }
            },

            'canned_tomatoes': {
                'minimum_sizes': {
                    'can': '14.5 oz can'
                },
                'conversions': {
                    'can': 14.5  # ounces per can
                }
            },
            'sugar': {
                'minimum_sizes': {
                    'teaspoon': '1 lb bag',
                    'tablespoon': '1 lb bag',
                    'cup': '2 lb bag',
                    'pound': '1 lb bag'
                },
                'conversions': {
                    'teaspoon': 0.0104167,  # pounds
                    'tablespoon': 0.0312500,  # pounds
                    'cup': 0.5,  # pounds
                }
            },
            'granulated_sugar': {
                'minimum_sizes': {
                    'teaspoon': '1 lb bag',
                    'tablespoon': '1 lb bag',
                    'cup': '2 lb bag',
                    'pound': '1 lb bag'
                },
                'conversions': {'cup': 0.5}  # pounds per cup
            },
            'brown_sugar': {
                'minimum_sizes': {
                    'teaspoon': '1 lb box',
                    'tablespoon': '1 lb box',
                    'cup': '2 lb box',
                    'pound': '1 lb box'
                },
                'conversions': {'cup': 0.5}  # pounds per cup
            },
            'powdered_sugar': {
                'minimum_sizes': {
                    'teaspoon': '1 lb box',
                    'tablespoon': '1 lb box',
                    'cup': '2 lb box',
                    'pound': '1 lb box'
                },
                'conversions': {'cup': 0.5}  # pounds per cup
            },
            'raisins': {
                'minimum_sizes': {
                    'teaspoon': '1.5 oz box',
                    'tablespoon': '1.5 oz box',
                    'cup': '15 oz box'
                },
                'conversions': {'cup': 5}  # ounces per cup
            },
            'all_purpose_flour': {
                'minimum_sizes': {
                    'teaspoon': '2 lb bag',
                    'tablespoon': '2 lb bag',
                    'cup': '5 lb bag',
                    'pound': '2 lb bag'
                },
                'conversions': {'cup': 0.4125}  # pounds per cup
            },
            'flour': {
                'minimum_sizes': {
                    'teaspoon': '1 lb bag',
                    'tablespoon': '1 lb bag',
                    'cup': '2 lb bag',
                    'pound': '1 lb bag'
                },
                'conversions': {
                    'teaspoon': 0.00859375,  # pounds
                    'tablespoon': 0.025781,  # pounds
                    'cup': 0.4125,  # pounds
                }
            },
            # Liquids
            'vegetable_broth': {
                'minimum_sizes': {
                    'teaspoon': '8 fl oz carton',
                    'tablespoon': '8 fl oz carton',
                    'cup': '32 fl oz carton',
                    'fluid_ounce': '8 fl oz carton'
                },
                'conversions': {
                    'teaspoon': 0.166667,  # fluid ounces
                    'tablespoon': 0.5,  # fluid ounces
                    'cup': 8,  # fluid ounces
                }
            },
            'olive_oil': {
                'minimum_sizes': {
                    'teaspoon': '8.5 fl oz bottle',
                    'tablespoon': '8.5 fl oz bottle',
                    'cup': '16 fl oz bottle',
                    'fluid_ounce': '8.5 fl oz bottle'
                },
                'conversions': {
                    'teaspoon': 0.166667,  # fluid ounces
                    'tablespoon': 0.5,  # fluid ounces
                    'cup': 8,  # fluid ounces
                }
            },
            # Spices & Seasonings
            'ground_cinnamon': {
                'minimum_sizes': {
                    'teaspoon': '1.5 oz jar',
                    'tablespoon': '1.5 oz jar',
                    'cup': '2.37 oz jar',
                    'ounce': '1.5 oz jar'
                },
                'conversions': {
                    'teaspoon': 0.0729167,  # ounces
                    'tablespoon': 0.21875,  # ounces
                    'cup': 3.5,  # ounces
                }
            },
            'vanilla_extract': {
                'minimum_sizes': {
                    'teaspoon': '1 fl oz bottle',
                    'tablespoon': '2 fl oz bottle',
                    'cup': '4 fl oz bottle',
                    'fluid_ounce': '1 fl oz bottle'
                },
                'conversions': {
                    'teaspoon': 0.166667,  # fluid ounces
                    'tablespoon': 0.5,  # fluid ounces
                    'cup': 8,  # fluid ounces
                }
            },
            # Dairy & Refrigerated
            'milk': {
                'minimum_sizes': {
                    'teaspoon': '8 fl oz bottle',
                    'tablespoon': '8 fl oz bottle',
                    'cup': '16 fl oz bottle',
                    'fluid_ounce': '8 fl oz bottle'
                },
                'conversions': {
                    'teaspoon': 0.166667,  # fluid ounces
                    'tablespoon': 0.5,  # fluid ounces
                    'cup': 8,  # fluid ounces
                }
            },
            'baking_powder': {
                'minimum_sizes': {
                    'teaspoon': '8 oz can',
                    'tablespoon': '8 oz can',
                    'cup': '8 oz can'
                },
                'conversions': {'cup': 8}  # ounces per cup
            },

            # Condiments
            'garlic': {
                'minimum_sizes': {
                    'clove': '1 bulb',
                },
                'conversions': {'clove': 1}  # Each clove is a single unit
            },
            'lemon_juice': {
                'minimum_sizes': {
                    'teaspoon': '4 fl oz bottle',
                    'tablespoon': '8 fl oz bottle',
                    'cup': '16 fl oz bottle'
                },
                'conversions': {'cup': 8}  # fluid ounces per cup
            },

            # Dairy
            'butter': {
                'minimum_sizes': {
                    'gram': '200g package',
                    'cup': '16 oz package (2 cups)',
                    'pound': '1 lb package'
                },
                'conversions': {'gram': 0.0353}  # pounds per gram
            },
            # Sweets
            'dark_chocolate': {
                'minimum_sizes': {
                    'gram': '200g bar',
                    'ounce': '7 oz bar',
                    'pound': '1 lb bag'
                },
                'conversions': {'gram': 0.0353}  # pounds per gram
            }
        }
        try:
            amount = float(amount)
        except (ValueError, TypeError):
            return f"{amount} {unit}", "Unable to determine package size"

        ingredient = ingredient_type.lower().strip()
        if ingredient in retail_sizes:
            data = retail_sizes[ingredient]
            conversion_factor = data['conversions'].get(unit, 1)  # Default to 1 if no conversion is defined
            converted_amount = amount * conversion_factor

            # Get recommended package size based on converted amount
            recommended_package = None
            for package_size, details in data['minimum_sizes'].items():
                if isinstance(details, dict) and converted_amount >= details.get('min_amount', 0):
                    recommended_package = details
                    break

                if recommended_package:
                    # Ensure consistent unit for package size recommendation
                    package_unit = recommended_package['unit']
                    if package_unit != unit and package_unit in data['conversions']:
                        package_amount = converted_amount / float(data['conversions'][package_unit])
                        return f"{package_amount:.2f} {package_unit}", recommended_package['description']
                if recommended_package is not None:
                    return converted_amount, recommended_package['description']
                else:
                    # Return a message or handle the case where no package is found
                    return converted_amount, "N/A (no suitable package size)"
            else:
                return converted_amount, "N/A (consider smaller unit)"  # Suggest using a smaller unit

        return f"{amount} {unit}", "Unable to determine package size"

    def setup_update_tab(self):
        """Setup the 'Update' tab with a search box, list of recipes_images, and update form."""
        # Create search box
        self.update_search_var = self.create_search_box(self.update_frame, self.update_recipe_list)

        # Create listbox with scrollbar
        list_frame = ttk.Frame(self.update_frame)
        list_frame.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Label(list_frame, text="Select ManageRecipe to Update:").pack(pady=5)

        # Add scrollbar
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")

        self.update_listbox = tk.Listbox(list_frame, width=50, height=10, yscrollcommand=scrollbar.set)
        self.update_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.update_listbox.yview)

        # Bind selection event
        self.update_listbox.bind('<<ListboxSelect>>', self.load_recipe_for_update)

        # Create update form frame
        self.update_form = ttk.Frame(self.update_frame)
        self.update_form.pack(pady=10, fill="both", expand=True)

        # Initial population of the list
        self.update_recipe_list()

    def update_recipe(self):
        """Update the selected recipe."""
        if not hasattr(self, 'current_recipe_index'):
            messagebox.showwarning("Warning", "Please select a recipe to update.")
            return

        # Get the selected recipe
        selection = self.update_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a recipe to update.")
            return

        recipe_name = self.update_listbox.get(selection[0])
        recipe_index = next((i for i, r in enumerate(self.recipes) if r['name'] == recipe_name), None)

        if recipe_index is None:
            messagebox.showerror("Error", "ManageRecipe not found.")
            return

        # Create updated recipe dictionary
        updated_recipe = {
            "name": self.name_var.get().strip(),
            "ingredients": [],
            "instructions": self.instructions_text.get("1.0", tk.END).strip().split("\n"),
            "image_path": self.image_path_var.get().strip()
        }

        # Add ingredients
        for vars in self.ingredient_vars:
            if vars[0].get().strip():  # Only add if ingredient name is not empty
                ingredient = {
                    "ingredient": vars[0].get().strip(),
                    "category": vars[1].get().strip(),
                    "quantity": vars[2].get().strip(),
                    "unit": vars[3].get().strip(),
                    "description": vars[4].get().strip()
                }
                updated_recipe["ingredients"].append(ingredient)

        # Update the recipe in the list
        self.recipes[recipe_index] = updated_recipe

        # Save to file
        try:
            self.save_recipes()
            messagebox.showinfo("Success", "ManageRecipe updated successfully!")
            self.refresh_lists()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save recipe: {str(e)}")

    def setup_delete_tab(self):
        """Setup the 'Delete' tab with a search box and list of recipes_images."""
        # Create search box
        self.delete_search_var = self.create_search_box(self.delete_frame, self.update_delete_list)

        # Create listbox with scrollbar
        list_frame = ttk.Frame(self.delete_frame)
        list_frame.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Label(list_frame, text="Select ManageRecipe to Delete:").pack(pady=5)

        # Add scrollbar
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")

        self.delete_listbox = tk.Listbox(list_frame, width=50, height=10, yscrollcommand=scrollbar.set)
        self.delete_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.delete_listbox.yview)

        # Add delete button
        ttk.Button(self.delete_frame, text="Delete Selected ManageRecipe",
                   command=self.delete_recipe).pack(pady=10)

        # Initial population of the list
        self.update_delete_list()

    def refresh_lists(self):
        """Refresh all Listboxes displaying recipes_images."""
        self.update_show_list()
        self.update_recipe_list()
        self.update_delete_list()

    def load_recipe_for_update(self, event=None):
        """Load the selected recipe details into the update form."""
        selection = self.update_listbox.curselection()
        if not selection:
            return

        # Clear previous form contents
        for widget in self.update_form.winfo_children():
            widget.destroy()

        # Create a main container frame
        main_container = ttk.Frame(self.update_form)
        main_container.pack(fill="both", expand=True)

        # Create scrollable area
        canvas = tk.Canvas(main_container)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        # Configure scrolling
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        # Create a separate frame for buttons that will stay at the bottom
        buttons_frame = ttk.Frame(self.update_form)
        buttons_frame.pack(side="bottom", fill="x", pady=10)

        # Pack the canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True, padx=5)
        scrollbar.pack(side="right", fill="y")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Get selected recipe
        recipe_index = selection[0]
        recipe = self.recipes[recipe_index]
        self.current_recipe_index = recipe_index

        # ManageRecipe name
        name_frame = ttk.Frame(scrollable_frame)
        name_frame.pack(fill="x", pady=5)
        ttk.Label(name_frame, text="ManageRecipe Name:").pack(side="left")
        self.name_var = tk.StringVar(value=recipe["name"])
        ttk.Entry(name_frame, textvariable=self.name_var, width=40).pack(side="left", padx=5)

        # Image section
        image_frame = ttk.LabelFrame(scrollable_frame, text="ManageRecipe Image")
        image_frame.pack(fill="x", pady=10, padx=5)

        # Image path entry and browse button
        image_path_frame = ttk.Frame(image_frame)
        image_path_frame.pack(fill="x", pady=5)

        self.image_path_var = tk.StringVar(value=recipe.get("image_path", ""))
        ttk.Label(image_path_frame, text="Image Path:").pack(side="left", padx=5)
        ttk.Entry(image_path_frame, textvariable=self.image_path_var, width=40).pack(side="left", padx=5)
        ttk.Button(image_path_frame, text="Browse", command=self.browse_image).pack(side="left", padx=5)

        # Image preview
        preview_frame = ttk.Frame(image_frame)
        preview_frame.pack(fill="x", pady=5)
        self.preview_label = ttk.Label(preview_frame, text="No image selected")
        self.preview_label.pack()

        # Update image preview if path exists
        if recipe.get("image_path") and os.path.exists(recipe.get("image_path")):
            try:
                image = Image.open(recipe["image_path"])
                image.thumbnail((200, 200))  # Resize image while maintaining aspect ratio
                photo = ImageTk.PhotoImage(image)
                self.preview_label.configure(image=photo)
                self.preview_label.image = photo  # Keep a reference
            except Exception as e:
                self.preview_label.configure(text=f"Error loading image: {str(e)}")

        # Ingredients section
        ingredients_frame = ttk.LabelFrame(scrollable_frame, text="Ingredients")
        ingredients_frame.pack(fill="x", pady=10, padx=5)

        # Headers for ingredients
        headers = ["Ingredient", "Category", "Quantity", "Unit", "Description"]
        for i, header in enumerate(headers):
            ttk.Label(ingredients_frame, text=header).grid(row=0, column=i, padx=5, pady=5)

        # Ingredient entries with delete buttons
        self.ingredient_vars = []
        for i, ing in enumerate(recipe.get("ingredients", [])):
            row_vars = []
            for j, field in enumerate(["ingredient", "category", "quantity", "unit", "description"]):
                var = tk.StringVar(value=ing.get(field, ""))
                entry = ttk.Entry(ingredients_frame, textvariable=var)
                entry.grid(row=i + 1, column=j, padx=2, pady=2)
                row_vars.append(var)

            # Add delete button for each ingredient row
            delete_btn = ttk.Button(ingredients_frame,
                                    text="Delete Ingredient",
                                    command=lambda row=i: self.delete_ingredient_row(row))
            delete_btn.grid(row=i + 1, column=len(headers), padx=2, pady=2)

            self.ingredient_vars.append(row_vars)

        # Ingredients management buttons
        ingredient_buttons_frame = ttk.Frame(ingredients_frame)
        ingredient_buttons_frame.grid(row=len(recipe.get("ingredients", [])) + 2,
                                      column=0,
                                      columnspan=len(headers) + 1,
                                      pady=5)

        ttk.Button(ingredient_buttons_frame,
                   text="Add Ingredient",
                   command=lambda: self.add_ingredient_row_update(ingredients_frame)).pack(side="left", padx=5)

        ttk.Button(ingredient_buttons_frame,
                   text="Undo Delete",
                   command=self.undo_delete_ingredient).pack(side="left", padx=5)

        ttk.Button(ingredient_buttons_frame,
                   text="Save Ingredients",
                   command=self.save_ingredients_changes).pack(side="left", padx=5)

        # Instructions section
        instructions_frame = ttk.LabelFrame(scrollable_frame, text="Instructions")
        instructions_frame.pack(fill="both", expand=True, pady=10, padx=5)

        self.instructions_text = tk.Text(instructions_frame, height=5, width=50)
        self.instructions_text.pack(fill="both", expand=True, pady=5, padx=5)
        self.instructions_text.insert("1.0", "\n".join(recipe.get("instructions", [])))

        # Move buttons to the separate bottom frame
        save_button = ttk.Button(buttons_frame, text="Save Changes",
                                 command=self.update_recipe)
        save_button.pack(side="left", padx=5)

        cancel_button = ttk.Button(buttons_frame, text="Cancel",
                                   command=self.refresh_lists)
        cancel_button.pack(side="left", padx=5)

        # Update canvas scroll region
        canvas.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))

    def browse_image(self):
        """Open file dialog to select an image"""
        file_path = filedialog.askopenfilename(
            filetypes=[
                ("Image files", "*.png *.jpg *.jpeg *.gif *.bmp"),
                ("All files", "*.*")
            ]
        )
        if file_path:
            self.image_path_var.set(file_path)
            self.update_image_preview()

    def update_image_preview(self):
        """Update the image preview"""
        image_path = self.image_path_var.get()
        if image_path and os.path.exists(image_path):
            try:
                image = Image.open(image_path)
                # Resize image while maintaining aspect ratio
                image.thumbnail((200, 200))
                photo = ImageTk.PhotoImage(image)
                self.preview_label.configure(image=photo)
                self.preview_label.image = photo  # Keep a reference
            except Exception as e:
                self.preview_label.configure(text=f"Error loading image: {str(e)}")
        else:
            self.preview_label.configure(text="No image selected")

    def create_search_box(self, parent, callback):
        """Create a search box with label and entry."""
        search_frame = ttk.Frame(parent)
        search_frame.pack(fill="x", padx=5, pady=5)

        ttk.Label(search_frame, text="Search:").pack(side="left")
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var)
        search_entry.pack(side="left", fill="x", expand=True, padx=5)

        # Bind the search variable to the callback
        search_var.trace("w", lambda *args: callback())

        return search_var

    def update_recipe_list(self, search_term=""):
        """Update the recipe list in the Update tab based on search term."""
        if hasattr(self, 'update_listbox'):
            self.update_listbox.delete(0, tk.END)
            search_term = self.update_search_var.get().lower() if hasattr(self, 'update_search_var') else ""

            for recipe in self.recipes:
                if search_term in recipe['name'].lower():
                    self.update_listbox.insert(tk.END, recipe['name'])

    def delete_ingredient_row(self, row_index):
        """Delete an ingredient row from the update form."""
        if 0 <= row_index < len(self.ingredient_vars):
            # Store the deleted ingredient for potential undo
            if not hasattr(self, 'deleted_ingredients'):
                self.deleted_ingredients = []
            self.deleted_ingredients.append(self.ingredient_vars[row_index])

            del self.ingredient_vars[row_index]
            self.refresh_ingredient_display()

    def undo_delete_ingredient(self):
        """Restore the last deleted ingredient."""
        if hasattr(self, 'deleted_ingredients') and self.deleted_ingredients:
            # Restore the last deleted ingredient
            last_deleted = self.deleted_ingredients.pop()
            self.ingredient_vars.append(last_deleted)
            self.refresh_ingredient_display()
            messagebox.showinfo("Success", "Last deleted ingredient restored!")
        else:
            messagebox.showinfo("Info", "No ingredients to restore!")

    def save_ingredients_changes(self):
        """Save changes made to ingredients."""
        if not hasattr(self, 'current_recipe_index'):
            messagebox.showwarning("Warning", "No recipe selected.")
            return

        # Get current recipe
        recipe = self.recipes[self.current_recipe_index]

        # Update ingredients list
        recipe['ingredients'] = []
        for vars in self.ingredient_vars:
            if vars[0].get().strip():  # Only add if ingredient name is not empty
                ingredient = {
                    "ingredient": vars[0].get().strip(),
                    "category": vars[1].get().strip(),
                    "quantity": vars[2].get().strip(),
                    "unit": vars[3].get().strip(),
                    "description": vars[4].get().strip()
                }
                recipe['ingredients'].append(ingredient)

        # Save changes
        try:
            self.save_recipes()
            messagebox.showinfo("Success", "Ingredients updated successfully!")
            self.refresh_ingredient_display()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save ingredients: {str(e)}")

    def refresh_ingredient_display(self):
        """Refresh the ingredient display in the update form."""
        if hasattr(self, 'current_recipe_index'):
            self.load_recipe_for_update(None)


class LoginWindow:
    def __init__(self, root, on_login_success):
        self.root = root
        self.on_login_success = on_login_success
        self.login_window = tk.Toplevel(root)
        self.login_window.title("Login")
        # Remove fixed geometry to allow dynamic sizing
        self.login_window.protocol("WM_DELETE_WINDOW", lambda: root.destroy())

        # Center the window on screen
        self.center_window()

        # Apply theme
        self.login_window.configure(bg=COLORS['light'])

        # Create styled frame
        login_frame = ttk.Frame(self.login_window, style='Main.TFrame')
        login_frame.pack(expand=True, fill='both', padx=20, pady=20)

        # Header - Changed text to be more generic
        ttk.Label(login_frame,
                  text="Welcome to ManageRecipe Manager",
                  style='Header.TLabel').pack(pady=(0, 20))

        # Username field
        username_frame = ttk.Frame(login_frame)
        username_frame.pack(fill='x', pady=5)
        ttk.Label(username_frame,
                  text="Username:",
                  font=FONTS['normal']).pack(side='left', padx=5)
        self.username_entry = ttk.Entry(username_frame,
                                        font=FONTS['normal'])
        self.username_entry.pack(side='left', expand=True, fill='x', padx=5)

        # Password field
        password_frame = ttk.Frame(login_frame)
        password_frame.pack(fill='x', pady=5)
        ttk.Label(password_frame,
                  text="Password:",
                  font=FONTS['normal']).pack(side='left', padx=5)
        self.password_entry = ttk.Entry(password_frame,
                                        show="*",
                                        font=FONTS['normal'])
        self.password_entry.pack(side='left', expand=True, fill='x', padx=5)

        # Login button
        button_frame = ttk.Frame(login_frame)
        button_frame.pack(pady=20)
        ttk.Button(button_frame,
                   text="Login",
                   style='Success.TButton',
                   command=self.authenticate).pack(pady=5)

        # Add info label
        ttk.Label(login_frame,
                  text="Enter your credentials to access the recipe manager",
                  font=FONTS['small'],
                  wraplength=300).pack(pady=5)

        # Bind Enter key to login
        self.username_entry.bind('<Return>', lambda e: self.password_entry.focus())
        self.password_entry.bind('<Return>', lambda e: self.authenticate())

        # Set initial focus
        self.username_entry.focus()

    def center_window(self):
        """Center the login window on the screen"""
        # Wait for window to be ready
        self.login_window.update_idletasks()

        # Get screen dimensions
        screen_width = self.login_window.winfo_screenwidth()
        screen_height = self.login_window.winfo_screenheight()

        # Calculate position
        window_width = 400  # Initial width
        window_height = 300  # Initial height
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2

        # Set window size and position
        self.login_window.geometry(f'{window_width}x{window_height}+{x}+{y}')

    def authenticate(self):
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()

        # Assuming verify_login returns True for valid credentials
        if verify_login(username, password):  # Assuming this function exists
            self.login_window.destroy()  # Close login window
            self.on_login_success(username)
            print(f"Login successful. Showing main GUI.   {username}")
        else:
            messagebox.showerror("Login Failed", "Invalid username or password")


if __name__ == '__main__':
    root = tk.Tk()
    root.withdraw()  # Hide the root window until login
    recipe_manager = RecipeManager(root)

    # Show login window
    LoginWindow(root, recipe_manager.on_login_success)

    root.mainloop()

"""
self.categories=["Vegetables", "Fruits", "Frozen", "Not Frozen", "Dairy", "Grains", "Meat", "Others"]    can you enlarge that list?
self.categories = [
    "Vegetables", 
    "Fruits", 
    "Frozen", 
    "Dairy", 
    "Grains", 
    "Meat", 
    "Poultry", 
    "Seafood", 
    "Baking Ingredients", 
    "Condiments", 
    "Spices", 
    "Herbs", 
    "Snacks", 
    "Beverages", 
    "Canned Goods", 
    "Oils & Vinegars", 
    "Nuts & Seeds", 
    "Legumes", 
    "Pasta", 
    "Rice", 
    "Bread & Bakery", 
    "Cheese", 
    "Sweets & Desserts", 
    "Breakfast Foods", 
    "Soups", 
    "Salad Dressings", 
    "Tea & Coffee", 
    "Juices", 
    "Ethnic Foods", 
    "Gluten-Free", 
    "Vegan", 
    "Organic", 
    "Non-GMO", 
    "Health Foods", 
    "Cleaning Supplies",  # for shopping lists
    "Others"
]"""
